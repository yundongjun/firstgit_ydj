from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active

ws.append(["번호","영어","수학"])
for i in range(1,11):
    ws.append([i,randint(0,100),randint(0,100)])

col_B = ws["B"]
# print(col_B)
# for cell in col_B:
#     print(cell.value)

col_range = ws["B:C"]
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

row_title = ws[1]
# for cell in row_title:
#     print(cell.value)

# row_range= ws[2:6]
# for rows in row_range:
#     for cell in rows:
#         print(cell.value,end=" ")
#     print()

# from openpyxl.utils.cell import coordinate_from_string

# row_range= ws[2:ws.max_row]
# for rows in row_range:
#     for cell in rows:
#         # print(cell.value,end=" ")
#         # print(cell.coordinate, end = " ")
#         xy = coordinate_from_string(cell.coordinate)
#         # print(xy,end=" ")
#         print(xy[0],end=" ")
#         print(xy[1],end=" ")
#     print()

# print(tuple(ws.rows))
# print(tuple(ws.columns))

# for row in tuple(ws.rows):
#     print(row[1].value)

# for column in tuple(ws.columns):
#     print(column[0].value)

# for row in ws.iter_rows():
#     print(row[1].value)

# for row in ws.iter_cols():
    # print(row[0].value)

# for row in ws.iter_rows(min_row=2,max_row=11,min_col=2,max_col=3):
#     # print(row[0].value,row[1].value)
#     print(row)

for col in ws.iter_cols(min_row=1,max_row=5,min_col=1,max_col=3):
    print(col)

wb.save("sample.xlsx")