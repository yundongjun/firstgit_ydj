from openpyxl.styles import Font, Border, Side ,PatternFill, Alignment
from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

a1 = ws["A1"]
b1 = ws["B1"]
c1 = ws["C1"]

ws.column_dimensions["A"].width = 5
ws.row_dimensions[1].height = 50

a1.font = Font(color="FF0000",italic=True,bold=True)
#색깔,살짝눞인 글꼴,글자 두께

b1.font = Font(color="CC33FF",name="Arial",strike=True)
#색깔,글꼴,취소선

c1.font = Font(color="0000FF",size=20,underline="single")
# 크기,밑줄

#테두리 적용
thin_border = Border(left=Side(style="thin"),right=Side(style="thin"),bottom=Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

for row in ws.rows:
    for cell in row:
        cell.aligment = Alignment(horizontal="center", vertical="center")
        if cell.column == 1:
            continue
        
        if isinstance(cell.value, int)and cell.value >90:
            cell.fill = PatternFill(fgColor="00FF00", fill_type = "solid")
            cell.font = Font(color="FF0000")

ws.freeze_panes = "B2"

wb.save("sample_style.xlsx")